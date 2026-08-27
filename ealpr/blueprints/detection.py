"""
blueprints/detection.py — License plate detection, video processing, and history.

Routes:
    POST /process_plate
    POST /process_video
    GET  /video_feed
    GET  /detection_history
    GET  /uploaded_images
    GET  /api/detection/<detection_id>
    POST /api/detections/clear
    GET  /api/export/detections
    GET  /api/images/<detection_id>
    GET  /api/activities
"""
import io
import os
import csv
import base64
from datetime import datetime, timedelta

from flask import Blueprint, render_template, request, jsonify, Response, flash, redirect, url_for
from flask_login import login_required, current_user
from mongoengine import Q

from ealpr.extensions import DB_ENABLED
from ealpr.decorators import db_required
from ealpr.image_storage import save_detection_images, encode_detection_image
from config import (
    ALLOWED_EXTENSIONS, TEMP_DIR,
    ALLOWED_VIDEO_EXTENSIONS, VIDEO_FRAME_SKIP, MAX_VIDEO_SIZE,
)
from models import Visitor, DetectionResult
from ealpr.ai_models import get_plate_detection_model, get_ocr_model
from ealpr.extensions import socketio
from ealpr.utils import (
    utc_to_cairo,
    process_license_plate,
    process_frame_and_detect,
    generate_frames,
)

detection_bp = Blueprint("detection", __name__)


# ── Image upload & plate detection ───────────────────────────────────────────

@detection_bp.route("/process_plate", methods=["POST"])
@login_required
def process_plate():
    plate_detection_model = get_plate_detection_model()
    ocr_model = get_ocr_model()
    if plate_detection_model is None or ocr_model is None:
        return jsonify({"error": "Models not loaded properly"}), 500
    if "image" not in request.files:
        return jsonify({"error": "No image provided"}), 400

    image = request.files["image"]
    if image.filename == "":
        return jsonify({"error": "No selected file"}), 400
    if not image.filename.lower().endswith(tuple(ALLOWED_EXTENSIONS)):
        return jsonify({
            "error": f'Unsupported file type. Allowed: {", ".join(ALLOWED_EXTENSIONS)}'
        }), 400

    os.makedirs(TEMP_DIR, exist_ok=True)
    temp_path = os.path.join(TEMP_DIR, image.filename)
    detection_record = None

    try:
        image.save(temp_path)
        processed_img, plate_text, success, error_msg, plate_confidence, ocr_confidence, char_details = process_license_plate(
            temp_path, plate_detection_model, ocr_model
        )

        if not success:
            if os.path.exists(temp_path):
                os.remove(temp_path)
            return jsonify({"success": False, "message": error_msg, "status": "error"})

        import cv2
        _, buffer = cv2.imencode(".jpg", processed_img)
        processed_img_base64 = base64.b64encode(buffer).decode("utf-8")

        visitor = None
        status = "unknown"
        visitor_name = None
        visitor_info = None

        if DB_ENABLED:
            try:
                visitor = Visitor.objects(license_plate=plate_text).first()
                status = "authorized" if visitor else "unauthorized"
                visitor_name = visitor.name if visitor else None

                # Auto-set entry time on first detection
                if visitor and not visitor.entry_datetime_utc:
                    try:
                        visitor.entry_datetime_utc = datetime.utcnow()
                        visitor.save()
                    except Exception as upd_err:
                        print(f"Error setting entry time for {visitor.name}: {upd_err}")

                if visitor:
                    entry_time_str = (
                        visitor.entry_time if isinstance(visitor.entry_time, str)
                        else utc_to_cairo(visitor.entry_time) if isinstance(visitor.entry_time, datetime)
                        else None
                    )
                    exit_time_str = (
                        visitor.exit_time if isinstance(visitor.exit_time, str)
                        else utc_to_cairo(visitor.exit_time) if isinstance(visitor.exit_time, datetime)
                        else None
                    )
                    visitor_info = {
                        "name": visitor.name,
                        "visitor_id": str(visitor.visitor_id),
                        "license_plate": visitor.license_plate,
                        "entry_time": entry_time_str,
                        "exit_time": exit_time_str,
                        "authorized": visitor.authorized,
                        "status": visitor.status,
                        "responsible_department": visitor.responsible_department,
                        "general_department": visitor.general_department,
                    }
            except Exception as db_err:
                print(f"Error querying visitor: {db_err}")

        detection_timestamp = datetime.utcnow()
        if DB_ENABLED:
            try:
                original_bytes = open(temp_path, "rb").read()
                processed_bytes = buffer.tobytes()
                original_path, processed_path = save_detection_images(original_bytes, processed_bytes)

                detection_record = DetectionResult(
                    plate_number=plate_text,
                    confidence=plate_confidence,
                    ocr_confidence=ocr_confidence,
                    status=status,
                    visitor_name=visitor_name,
                    processed_by=current_user if current_user.is_authenticated else None,
                    original_image_path=original_path,
                    processed_image_path=processed_path,
                    timestamp=detection_timestamp,
                )
                detection_record.save()
            except Exception as db_err:
                print(f"Error saving detection record: {db_err}")

        display_confidence = float(plate_confidence) if isinstance(plate_confidence, (int, float)) else 0.0

        return jsonify({
            "success": True,
            "plate_text": plate_text,
            "status": status,
            "visitor_name": visitor_name,
            "processed_image": processed_img_base64,
            "confidence": display_confidence,
            "ocr_confidence": ocr_confidence,
            "char_details": char_details,
            "visitor_id": str(visitor.id) if visitor else None,
            "visitor_info": visitor_info,
            "detection_timestamp": utc_to_cairo(detection_timestamp) if detection_timestamp else None,
        })

    except Exception as e:
        print(f"Error processing image: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        try:
            if os.path.exists(temp_path):
                os.remove(temp_path)
        except Exception as cleanup_err:
            print(f"Error cleaning up temp file: {cleanup_err}")


# ── Video upload & batch processing ─────────────────────────────────────────

@detection_bp.route("/process_video", methods=["POST"])
@login_required
def process_video():
    plate_detection_model = get_plate_detection_model()
    ocr_model = get_ocr_model()
    if plate_detection_model is None or ocr_model is None:
        return jsonify({"error": "Models not loaded properly"}), 500
    if "video" not in request.files:
        return jsonify({"error": "No video file provided"}), 400

    video = request.files["video"]
    if video.filename == "":
        return jsonify({"error": "No selected file"}), 400

    file_ext = video.filename.rsplit(".", 1)[1].lower() if "." in video.filename else ""
    if file_ext not in ALLOWED_VIDEO_EXTENSIONS:
        return jsonify({
            "error": f'Unsupported file type. Allowed: {", ".join(ALLOWED_VIDEO_EXTENSIONS)}'
        }), 400

    video.seek(0, os.SEEK_END)
    file_size = video.tell()
    video.seek(0)
    if file_size > MAX_VIDEO_SIZE:
        return jsonify({
            "error": f"Video file size exceeds limit of {MAX_VIDEO_SIZE / 1024 / 1024:.2f} MB"
        }), 400

    import cv2
    os.makedirs(TEMP_DIR, exist_ok=True)
    temp_path = os.path.join(TEMP_DIR, video.filename)
    output_path = os.path.join(TEMP_DIR, f"processed_{video.filename}")

    try:
        video.save(temp_path)
        cap = cv2.VideoCapture(temp_path)
        if not cap.isOpened():
            raise Exception("Could not open video file")

        fps = cap.get(cv2.CAP_PROP_FPS)
        frame_width = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
        frame_height = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
        total_frames = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))

        fourcc = cv2.VideoWriter_fourcc(*"mp4v")
        out = cv2.VideoWriter(output_path, fourcc, fps, (frame_width, frame_height))

        frame_count = 0
        processed_frames = 0
        best_detections = {}
        last_frame = None

        while True:
            ret, frame = cap.read()
            if not ret:
                break

            frame_count += 1
            if frame_count % VIDEO_FRAME_SKIP == 0:
                detection_data, processed_img, current_frame = process_frame_and_detect(
                    frame, frame_count, plate_detection_model, ocr_model, None, last_frame
                )
                if detection_data:
                    plate_text = detection_data["plate_text"]
                    if (plate_text not in best_detections or
                            detection_data["confidence"] > best_detections[plate_text]["confidence"]):
                        best_detections[plate_text] = detection_data
                        try:
                            socketio.emit("live_detection", detection_data)
                        except Exception as emit_err:
                            print(f"Error emitting detection: {emit_err}")

                out.write(processed_img if processed_img is not None else frame)
                processed_frames += 1
                last_frame = current_frame

        cap.release()
        out.release()

        video_base64 = None
        if os.path.exists(output_path):
            with open(output_path, "rb") as vf:
                video_base64 = base64.b64encode(vf.read()).decode("utf-8")

        for path in [temp_path, output_path]:
            if os.path.exists(path):
                os.remove(path)

        return jsonify({
            "success": True,
            "message": "Video processing completed",
            "total_frames": total_frames,
            "processed_frames": processed_frames,
            "detections": list(best_detections.values()),
            "processed_video": video_base64,
        })

    except Exception as e:
        for path in [temp_path, output_path]:
            if os.path.exists(path):
                try:
                    os.remove(path)
                except Exception:
                    pass
        return jsonify({"error": f"An error occurred during video processing: {e}"}), 500


# ── Live camera MJPEG stream ─────────────────────────────────────────────────

@detection_bp.route("/video_feed")
@login_required
def video_feed():
    plate_detection_model = get_plate_detection_model()
    ocr_model = get_ocr_model()
    if plate_detection_model is None or ocr_model is None:
        return Response("Models not loaded properly", status=500)
    return Response(
        generate_frames(),
        mimetype="multipart/x-mixed-replace; boundary=frame"
    )


# ── Detection history (paginated) ────────────────────────────────────────────

@detection_bp.route("/detection_history")
@login_required
@db_required
def detection_history():
    page = request.args.get("page", 1, type=int)
    per_page = 10
    skip = (page - 1) * per_page
    total = DetectionResult.objects.count()
    detections = DetectionResult.objects.order_by("-timestamp")[skip: skip + per_page]
    total_pages = (total + per_page - 1) // per_page
    return render_template(
        "detection_history.html",
        detections=detections,
        page=page,
        per_page=per_page,
        total=total,
        total_pages=total_pages,
        max=max,
        min=min,
    )


# ── Uploaded images gallery ──────────────────────────────────────────────────

@detection_bp.route("/uploaded_images")
@login_required
@db_required
def uploaded_images():
    try:
        detections = DetectionResult.objects.order_by("-timestamp").all()
        return render_template("uploaded_images.html", detections=detections)
    except Exception as e:
        from flask import flash
        print(f"Error retrieving detections: {e}")
        flash("An error occurred while loading images. Please try again.", "error")
        return render_template("uploaded_images.html", detections=[])


# ── Single detection detail (JSON) ──────────────────────────────────────────

@detection_bp.route("/api/detection/<string:detection_id>")
@login_required
@db_required
def get_detection(detection_id):
    from bson.objectid import InvalidId
    try:
        detection = DetectionResult.objects(id=detection_id).first()
        if not detection:
            return jsonify({"success": False, "message": "Detection record not found"}), 404

        processed_by_username = detection.processed_by.username if detection.processed_by else None
        visitor = Visitor.objects(license_plate=detection.plate_number).first() if detection.plate_number else None

        if visitor:
            entry_time_str = (
                visitor.entry_time if isinstance(visitor.entry_time, str)
                else utc_to_cairo(visitor.entry_time) if isinstance(visitor.entry_time, datetime)
                else None
            )
            exit_time_str = (
                visitor.exit_time if isinstance(visitor.exit_time, str)
                else utc_to_cairo(visitor.exit_time) if isinstance(visitor.exit_time, datetime)
                else None
            )
            visitor_info = {
                "name": visitor.name or "N/A",
                "visitor_id": str(visitor.visitor_id) if visitor.visitor_id else "N/A",
                "license_plate": visitor.license_plate,
                "entry_time": entry_time_str or "N/A",
                "exit_time": exit_time_str or "N/A",
                "department": visitor.responsible_department or "N/A",
            }
        else:
            visitor_info = {
                "name": detection.visitor_name or "N/A",
                "visitor_id": "N/A",
                "license_plate": detection.plate_number,
                "entry_time": "N/A",
                "exit_time": "N/A",
                "department": "N/A",
            }

        timestamp_str = (
            detection.timestamp if isinstance(detection.timestamp, str)
            else utc_to_cairo(detection.timestamp) if isinstance(detection.timestamp, datetime)
            else None
        )

        return jsonify({
            "success": True,
            "id": str(detection.id),
            "timestamp": timestamp_str,
            "plate_number": detection.plate_number,
            "confidence": float(detection.confidence) if detection.confidence is not None else 0.0,
            "status": detection.status,
            "visitor_name": detection.visitor_name,
            "processed_by": processed_by_username,
            "original_image": encode_detection_image(detection.original_image_path, detection.original_image),
            "processed_image": encode_detection_image(detection.processed_image_path, detection.processed_image),
            "visitor_info": visitor_info,
        })
    except Exception as e:
        print(f"Error fetching detection {detection_id}: {e}")
        return jsonify({"success": False, "message": f"Error fetching detection details: {e}"}), 500


# ── Clear all detection history (admin) ─────────────────────────────────────

@detection_bp.route("/api/detections/clear", methods=["POST"])
@login_required
@db_required
def clear_detection_history():
    if not current_user.is_admin:
        return jsonify({"success": False, "message": "Access denied"}), 403
    try:
        DetectionResult.objects.delete()
        return jsonify({"success": True, "message": "Detection history cleared successfully"})
    except Exception as e:
        return jsonify({"success": False, "message": f"Failed to clear detection history: {e}"}), 500


# ── Export detections as CSV or Excel ────────────────────────────────────────

@detection_bp.route("/api/export/detections")
@login_required
@db_required
def export_detections():
    try:
        status_filter = request.args.get("status", "all")
        date_filter = request.args.get("date_filter", "all")
        fmt = request.args.get("format", "csv")

        query = DetectionResult.objects
        if status_filter != "all":
            query = query.filter(status=status_filter)
        if date_filter != "all":
            now = datetime.utcnow()
            if date_filter == "today":
                query = query.filter(timestamp__gte=now.replace(hour=0, minute=0, second=0))
            elif date_filter == "week":
                query = query.filter(timestamp__gte=now - timedelta(days=7))
            elif date_filter == "month":
                query = query.filter(timestamp__gte=now - timedelta(days=30))

        results = query.order_by("-timestamp")
        if not results:
            return jsonify({"error": "No data found for the selected filters"}), 404

        if fmt == "csv":
            output = io.StringIO()
            writer = csv.writer(output)
            writer.writerow(["Timestamp", "Plate Number", "Status", "Confidence", "Visitor Name", "Processed By"])
            for r in results:
                processed_by_username = r.processed_by.username if r.processed_by else "N/A"
                writer.writerow([
                    utc_to_cairo(r.timestamp),
                    r.plate_number,
                    r.status,
                    f"{r.confidence:.2f}%" if r.confidence else "N/A",
                    r.visitor_name or "N/A",
                    processed_by_username,
                ])
            output.seek(0)
            return Response(
                output,
                mimetype="text/csv",
                headers={"Content-Disposition": f'attachment; filename=detections_export_{datetime.utcnow().strftime("%Y%m%d")}.csv'},
            )

        elif fmt == "excel":
            from openpyxl import Workbook

            wb = Workbook()
            ws = wb.active
            ws.title = "Detection History"
            headers = ["Timestamp", "Plate Number", "Status", "Confidence", "Visitor Name", "Processed By"]
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            for row_idx, r in enumerate(results, 2):
                processed_by_username = r.processed_by.username if r.processed_by else "N/A"
                ws.cell(row=row_idx, column=1, value=utc_to_cairo(r.timestamp))
                ws.cell(row=row_idx, column=2, value=r.plate_number)
                ws.cell(row=row_idx, column=3, value=r.status)
                ws.cell(row=row_idx, column=4, value=f"{r.confidence:.2f}%" if r.confidence else "N/A")
                ws.cell(row=row_idx, column=5, value=r.visitor_name or "N/A")
                ws.cell(row=row_idx, column=6, value=processed_by_username)
            excel_file = io.BytesIO()
            wb.save(excel_file)
            excel_file.seek(0)
            return Response(
                excel_file,
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f'attachment; filename=detections_export_{datetime.utcnow().strftime("%Y%m%d")}.xlsx'},
            )
        else:
            return jsonify({"error": "Unsupported export format"}), 400

    except Exception as e:
        return jsonify({"error": f"Failed to export data: {e}"}), 500


# ── Single detection image retrieval ─────────────────────────────────────────

@detection_bp.route("/api/images/<string:detection_id>")
@login_required
@db_required
def get_detection_image(detection_id):
    from bson.objectid import InvalidId
    try:
        detection = DetectionResult.objects(id=detection_id).first()
    except InvalidId:
        return jsonify({"success": False, "message": "Invalid detection ID format."}), 400

    if not detection:
        return jsonify({"success": False, "message": "Detection record not found."}), 404

    return jsonify({
        "success": True,
        "original_image": base64.b64encode(detection.original_image).decode("utf-8") if detection.original_image else "",
        "processed_image": base64.b64encode(detection.processed_image).decode("utf-8") if detection.processed_image else "",
        "plate_number": detection.plate_number,
        "status": detection.status,
        "timestamp": utc_to_cairo(detection.timestamp) if detection.timestamp else None,
        "visitor_name": detection.visitor_name,
    })


# ── Recent activity feed (last 10 detections) ─────────────────────────────

@detection_bp.route("/api/activities")
@login_required
@db_required
def get_activities():
    try:
        activities = DetectionResult.objects.order_by("-timestamp").limit(10)
        activity_list = [
            {
                "action": "detect_plate",
                "details": f"Detected plate {a.plate_number} - {a.status}",
                "timestamp": utc_to_cairo(a.timestamp),
            }
            for a in activities
        ]
        return jsonify({"success": True, "activities": activity_list})
    except Exception as e:
        return jsonify({"success": False, "error": "Failed to fetch activities"}), 500


# ── Plate number search API ───────────────────────────────────────────────

@detection_bp.route("/api/search/plate")
@login_required
@db_required
def search_plate():
    query = request.args.get("q", "").strip()
    if not query:
        return jsonify({"success": False, "message": "Search query required"}), 400
    try:
        # Search visitor database
        visitor = Visitor.objects(Q(license_plate__icontains=query) | Q(name__icontains=query)).first()
        if visitor:
            return jsonify({
                "success": True,
                "data": {
                    "id": str(visitor.id),
                    "name": visitor.name,
                    "license_plate": visitor.license_plate,
                    "status": "authorized" if visitor.authorized else "unauthorized"
                }
            })
        return jsonify({"success": False, "message": "No visitor found"}), 404
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500


# ── SocketIO connection events ───────────────────────────────────────────────


@socketio.on("connect")
def handle_connect():
    print("Client connected to SocketIO.")


@socketio.on("disconnect")
def handle_disconnect():
    print("Client disconnected from SocketIO.")
