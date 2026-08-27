"""
blueprints/reports.py — Report generation and analytics routes.

Routes:
    GET /reports
    GET /api/analytics/visitor-stats
    GET /api/analytics/export-excel
    GET /api/reports/<type>         (type: daily | weekly | monthly)
"""
import io
from datetime import datetime, timedelta

from flask import Blueprint, render_template, request, jsonify, Response, flash, redirect, url_for
from flask_login import login_required
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from ealpr.extensions import DB_ENABLED
from ealpr.decorators import db_required

from models import Visitor
from ealpr.utils import utc_to_cairo

reports_bp = Blueprint("reports", __name__)


@reports_bp.before_request
def check_db_available():
    if not DB_ENABLED:
        if request.path.startswith("/api/") or request.is_json:
            return jsonify({
                "success": False,
                "message": "Database unavailable. Please try again later."
            }), 503
        flash("Database is offline. Reports are currently unavailable.", "warning")
        return redirect(url_for("main.index"))


@reports_bp.route("/reports")
@login_required
@db_required
def reports():
    return render_template("reports.html")


# ── Visitor statistics (JSON) ────────────────────────────────────────────────

@reports_bp.route("/api/analytics/visitor-stats")
@login_required
@db_required
def get_visitor_stats():
    try:
        end_date = datetime.now()
        month_start = end_date - timedelta(days=30)

        total_visitors = Visitor.objects.count()
        authorized_visitors = Visitor.objects(status="authorized").count()
        unauthorized_visitors = Visitor.objects(status="unauthorized").count()
        pending_visitors = Visitor.objects(status="pending").count()

        # Last 7 days daily breakdown
        daily_traffic = []
        daily_status = []
        for i in range(7):
            day = end_date - timedelta(days=i)
            day_start = day.replace(hour=0, minute=0, second=0)
            day_end = day.replace(hour=23, minute=59, second=59)

            count = Visitor.objects(entry_datetime_utc__gte=day_start, entry_datetime_utc__lt=day_end).count()
            auth = Visitor.objects(entry_datetime_utc__gte=day_start, entry_datetime_utc__lt=day_end, status="authorized").count()
            unauth = Visitor.objects(entry_datetime_utc__gte=day_start, entry_datetime_utc__lt=day_end, status="unauthorized").count()
            pending = Visitor.objects(entry_datetime_utc__gte=day_start, entry_datetime_utc__lt=day_end, status="pending").count()

            daily_traffic.append({"date": day.strftime("%Y-%m-%d"), "day": day.strftime("%a"), "count": count})
            daily_status.append({"date": day.strftime("%Y-%m-%d"), "authorized": auth, "unauthorized": unauth, "pending": pending})

        daily_traffic.reverse()
        daily_status.reverse()

        # Peak hours (last 30 days)
        hour_counts = {}
        for hour in range(24):
            hour_key = f"{hour:02d}:00"
            hour_counts[hour_key] = {
                "total": Visitor.objects(entry_datetime_utc__gte=month_start, entry_time__startswith=f"{hour:02d}:").count(),
                "authorized": Visitor.objects(entry_datetime_utc__gte=month_start, entry_time__startswith=f"{hour:02d}:", status="authorized").count(),
                "unauthorized": Visitor.objects(entry_datetime_utc__gte=month_start, entry_time__startswith=f"{hour:02d}:", status="unauthorized").count(),
                "pending": Visitor.objects(entry_datetime_utc__gte=month_start, entry_time__startswith=f"{hour:02d}:", status="pending").count(),
            }

        # Department stats
        dept_stats = {}
        departments = Visitor.objects.distinct("responsible_department")
        for dept in departments:
            if dept:
                total = Visitor.objects(responsible_department=dept).count()
                dept_stats[dept] = {
                    "total": total,
                    "authorized": Visitor.objects(responsible_department=dept, status="authorized").count(),
                    "unauthorized": Visitor.objects(responsible_department=dept, status="unauthorized").count(),
                    "pending": Visitor.objects(responsible_department=dept, status="pending").count(),
                }

        return jsonify({
            "success": True,
            "summary": {
                "total_visitors": total_visitors,
                "authorized_visitors": authorized_visitors,
                "unauthorized_visitors": unauthorized_visitors,
                "pending_visitors": pending_visitors,
                "authorization_rate": (authorized_visitors / total_visitors * 100) if total_visitors > 0 else 0,
            },
            "daily_traffic": daily_traffic,
            "daily_status": daily_status,
            "peak_hours": hour_counts,
            "department_stats": dept_stats,
        })
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500


# ── Export analytics to Excel ────────────────────────────────────────────────

@reports_bp.route("/api/analytics/export-excel")
@login_required
@db_required
def export_analytics_excel():
    try:
        wb = Workbook()

        # Summary sheet
        ws_summary = wb.active
        ws_summary.title = "Summary"
        ws_summary.append(["Metric", "Value"])

        total_visitors = Visitor.objects.count()
        authorized_visitors = Visitor.objects(status="authorized").count()
        unauthorized_visitors = Visitor.objects(status="unauthorized").count()
        pending_visitors = Visitor.objects(status="pending").count()

        for row in [
            ["Total Visitors", total_visitors],
            ["Authorized Visitors", authorized_visitors],
            ["Unauthorized Visitors", unauthorized_visitors],
            ["Pending Visitors", pending_visitors],
            ["Authorization Rate", f"{(authorized_visitors / total_visitors * 100) if total_visitors > 0 else 0:.1f}%"],
        ]:
            ws_summary.append(row)
        for cell in ws_summary[1]:
            cell.font = Font(bold=True)
        for col in ["A", "B"]:
            ws_summary.column_dimensions[col].width = 20

        # Daily traffic sheet (last 30 days)
        ws_daily = wb.create_sheet("Daily Traffic")
        ws_daily.append(["Date", "Day", "Total", "Authorized", "Unauthorized", "Pending"])
        end_date = datetime.now()
        for i in range(30):
            day = end_date - timedelta(days=i)
            day_start = day.replace(hour=0, minute=0, second=0)
            day_end = day.replace(hour=23, minute=59, second=59)
            ws_daily.append([
                day.strftime("%Y-%m-%d"),
                day.strftime("%a"),
                Visitor.objects(entry_datetime_utc__gte=day_start, entry_datetime_utc__lt=day_end).count(),
                Visitor.objects(entry_datetime_utc__gte=day_start, entry_datetime_utc__lt=day_end, status="authorized").count(),
                Visitor.objects(entry_datetime_utc__gte=day_start, entry_datetime_utc__lt=day_end, status="unauthorized").count(),
                Visitor.objects(entry_datetime_utc__gte=day_start, entry_datetime_utc__lt=day_end, status="pending").count(),
            ])
        for cell in ws_daily[1]:
            cell.font = Font(bold=True)
        for col in ["A", "B", "C", "D", "E", "F"]:
            ws_daily.column_dimensions[col].width = 15

        # Hourly traffic sheet
        ws_hourly = wb.create_sheet("Hourly Traffic")
        ws_hourly.append(["Hour", "Total", "Authorized", "Unauthorized", "Pending"])
        month_start = end_date - timedelta(days=30)
        for hour in range(24):
            hour_str = f"{hour:02d}:00"
            ws_hourly.append([
                hour_str,
                Visitor.objects(entry_datetime_utc__gte=month_start, entry_time__startswith=f"{hour:02d}:").count(),
                Visitor.objects(entry_datetime_utc__gte=month_start, entry_time__startswith=f"{hour:02d}:", status="authorized").count(),
                Visitor.objects(entry_datetime_utc__gte=month_start, entry_time__startswith=f"{hour:02d}:", status="unauthorized").count(),
                Visitor.objects(entry_datetime_utc__gte=month_start, entry_time__startswith=f"{hour:02d}:", status="pending").count(),
            ])
        for cell in ws_hourly[1]:
            cell.font = Font(bold=True)
        for col in ["A", "B", "C", "D", "E"]:
            ws_hourly.column_dimensions[col].width = 15

        # Department statistics sheet
        ws_dept = wb.create_sheet("Department Statistics")
        ws_dept.append(["Department", "Total Visitors", "Authorized", "Unauthorized", "Pending", "Authorization Rate"])
        for dept in Visitor.objects.distinct("responsible_department"):
            if dept:
                total = Visitor.objects(responsible_department=dept).count()
                auth = Visitor.objects(responsible_department=dept, status="authorized").count()
                unauth = Visitor.objects(responsible_department=dept, status="unauthorized").count()
                pending = Visitor.objects(responsible_department=dept, status="pending").count()
                ws_dept.append([dept, total, auth, unauth, pending, f"{(auth / total * 100) if total > 0 else 0:.1f}%"])
        for cell in ws_dept[1]:
            cell.font = Font(bold=True)
        for col in ["A", "B", "C", "D", "E", "F"]:
            ws_dept.column_dimensions[col].width = 20

        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        return Response(
            excel_file.getvalue(),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename=visitor_analytics_{end_date.strftime("%Y%m%d")}.xlsx'},
        )
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500


# ── Styled period reports ────────────────────────────────────────────────────

@reports_bp.route("/api/reports/<type>")
@login_required
@db_required
def export_report(type):
    if type not in ["daily", "weekly", "monthly"]:
        return jsonify({"success": False, "message": "Invalid report type"}), 400

    excel_file = None
    try:
        wb = Workbook()
        ws = wb.active

        title_font = Font(bold=True, size=14, color="366092")
        header_font = Font(bold=True, color="FFFFFF")
        regular_font = Font(size=10)
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

        end_date = datetime.now()
        if type == "daily":
            start_date = end_date.replace(hour=0, minute=0, second=0, microsecond=0)
            title = f"Daily Visitor Report - {end_date.strftime('%Y-%m-%d')}"
        elif type == "weekly":
            start_date = end_date - timedelta(days=7)
            title = f"Weekly Visitor Report ({start_date.strftime('%Y-%m-%d')} - {end_date.strftime('%Y-%m-%d')})"
        else:
            start_date = end_date.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            title = f"Monthly Visitor Report - {end_date.strftime('%B %Y')}"

        visitors = Visitor.objects(
            entry_datetime_utc__gte=start_date,
            entry_datetime_utc__lte=end_date,
        ).order_by("entry_datetime_utc")

        ws.merge_cells("A1:G1")
        ws["A1"] = title
        ws["A1"].font = title_font
        ws["A1"].alignment = Alignment(horizontal="center")

        headers = ["Visitor Name", "ID", "License Plate", "Date", "Time", "Status", "Department"]
        ws.append([""] * len(headers))
        ws.append(headers)

        header_row = 3
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=header_row, column=col)
            cell.font = header_font
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for visitor in visitors:
            entry_time = utc_to_cairo(visitor.entry_datetime_utc)
            date_str, time_str = entry_time.split(" ")
            row = [
                visitor.name,
                visitor.visitor_code,
                visitor.license_plate,
                date_str,
                time_str,
                visitor.status.title(),
                visitor.responsible_department or "N/A",
            ]
            ws.append(row)
            current_row = ws.max_row
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=current_row, column=col)
                cell.font = regular_font
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="left")
                if current_row % 2 == 0:
                    cell.fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

        column_widths = {"A": 30, "B": 15, "C": 15, "D": 12, "E": 10, "F": 15, "G": 20}
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        ws.auto_filter.ref = f"A3:G{ws.max_row}"

        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        filename = f"{type}_report_{end_date.strftime('%Y%m%d')}.xlsx"
        return Response(
            excel_file.getvalue(),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )
    except Exception as e:
        if excel_file:
            try:
                excel_file.close()
            except Exception:
                pass
        return jsonify({"success": False, "message": str(e)}), 500
