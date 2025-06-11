import io
from flask import Flask, render_template, request, jsonify, Response, redirect, url_for, flash
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from datetime import datetime, timedelta
from zoneinfo import ZoneInfo  # Python 3.9+
import cv2
import numpy as np
from ultralytics import YOLO
import os
from dotenv import load_dotenv
import threading
import time
from config import (
    MODEL_PATH, CONFIDENCE_THRESHOLD,
    PLATE_CLASS_ID, OCR_MODEL_PATH,
    ALLOWED_EXTENSIONS, MAX_IMAGE_SIZE, TEMP_DIR,
    MONGODB_URI, MONGODB_DB_NAME, SECRET_KEY, TESSERACT_CMD,
    ALLOWED_VIDEO_EXTENSIONS, VIDEO_FRAME_SKIP, VIDEO_DUPLICATE_WINDOW, MAX_VIDEO_SIZE
)
from PIL import Image, ImageFont, ImageDraw
import base64
import re
import arabic_reshaper
from bidi.algorithm import get_display
from tensorflow import keras
from models import User, Visitor, UserActivity, SystemSettings, DetectionResult
from mongoengine import connect
import csv
from io import StringIO
from openpyxl import Workbook
from mongoengine.queryset.visitor import Q
from flask_socketio import SocketIO, emit
from bson.objectid import ObjectId, InvalidId

# Define class labels mapping from ocr_production.py
CLASS_LABELS_MAPPING = {
    0: "٠", 1: "١", 2: "٢", 3: "٣", 4: "٤", 5: "٥", 6: "٦", 7: "٧",
    8: "ح", 9: "٨", 10: "٩", 11: "ط", 12: "ظ", 13: "ع", 14: "أ", 15: "ب",
    16: "ض", 17: "د", 18: "ف", 19: "غ", 20: "ه", 21: "ج", 22: "ك", 23: "خ",
    24: "ل", 25: "م", 26: "ن", 27: "ق", 28: "ر", 29: "ص", 30: "س", 31: "ش",
    32: "ت", 33: "ث", 34: "و", 35: "ي", 36: "ذ", 37: "ز"
}

# Define colors for visualization
COLORS = [
    (255, 0, 0), (34, 75, 12), (0, 0, 255), (255, 255, 0),
    (255, 0, 255), (21, 52, 72), (66, 50, 168)
]

# Load environment variables
load_dotenv()

# Initialize Flask app
app = Flask(__name__)
app.config['SECRET_KEY'] = SECRET_KEY

# Initialize SocketIO
socketio = SocketIO(app)

# Connect to MongoDB
print("Connecting to MongoDB...")
connect(MONGODB_DB_NAME, host=MONGODB_URI)
print("MongoDB connection test:", Visitor.objects.count())  # Test connection
print("Connected to MongoDB successfully!")

# Configure Flask-Login
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'
login_manager.login_message = 'Please log in to access this page.'
login_manager.login_message_category = 'info'

# Load the models
try:
    print("Loading models...")
    plate_detection_model = YOLO(MODEL_PATH)
    ocr_model = YOLO(OCR_MODEL_PATH)
    print("Models loaded successfully")
except Exception as e:
    print(f"Error loading models: {str(e)}")
    plate_detection_model = None
    ocr_model = None

# Create necessary directories
os.makedirs(TEMP_DIR, exist_ok=True)

# Helper function to convert UTC to Cairo time
def utc_to_cairo(utc_time):
    if isinstance(utc_time, str):
        try:
            utc_time = datetime.fromisoformat(utc_time)
        except ValueError:
            return "Invalid date format"
    if utc_time.tzinfo is None:
        utc_time = utc_time.replace(tzinfo=ZoneInfo("UTC"))
    local_time = utc_time.astimezone(ZoneInfo("Africa/Cairo"))
    return local_time.strftime('%Y-%m-%d %H:%M:%S')

# Register Jinja2 filter for templates
@app.template_filter('to_cairo')
def to_cairo(value):
    return utc_to_cairo(value)

# Initialize default admin user
def create_default_admin():
    admin = User.objects(username='admin').first()
    if not admin:
        admin = User(
            username='admin',
            is_admin=True,
            created_at=datetime.utcnow()
        )
        admin.set_password('admin123')
        admin.save()
        print("Created default admin user")

# Create default admin on startup
create_default_admin()

# Define font path
font_path = r"D:\EALPR SYSTEM\alfont_com_arial-1.ttf"

def draw_arabic_text(image, text, position, color):
    """
    Draw Arabic text on an image with background rectangle
    """
    reshaped_text = arabic_reshaper.reshape(text)
    bidi_text = get_display(reshaped_text)
    img_pil = Image.fromarray(image)
    draw = ImageDraw.Draw(img_pil)
    font = ImageFont.truetype(font_path, 40)
    bbox = draw.textbbox(position, bidi_text, font=font)
    draw.rectangle(bbox, fill=color)
    draw.text(position, bidi_text, font=font, fill="black")
    return np.array(img_pil)

def reverse_arabic(text):
    """
    Reverse Arabic text while preserving numbers and adding spaces between characters
    """
    segments = []
    current_segment = ""
    is_number = text[0].isdigit() if text else False
    for char in text:
        if char.isdigit() == is_number:
            current_segment += char
        else:
            if current_segment:
                segments.append(current_segment)
            current_segment = char
            is_number = not is_number
    if current_segment:
        segments.append(current_segment)
    reversed_text = ""
    for i, segment in enumerate(segments):
        if segment[0].isdigit():
            # Add space before and after numbers
            if i > 0:
                reversed_text += " "
            reversed_text += segment
            if i < len(segments) - 1:
                reversed_text += " "
        else:
            # Add space between each Arabic character
            reversed_text = " ".join(segment[::-1]) + reversed_text
    return reversed_text.strip()  # Remove any extra spaces at the beginning or end

@login_manager.user_loader
def load_user(user_id):
    try:
        return User.objects(id=user_id).first()
    except:
        return None

# Define the b64encode filter for Jinja2
@app.template_filter('b64encode')
def b64encode_filter(data):
    if data is None:
        return ''
    return base64.b64encode(data).decode('utf-8')

@app.route('/')
@login_required
def index():
    total_detections = DetectionResult.objects.count()
    authorized_visitors = Visitor.objects(authorized=True).count()
    unauthorized_visitors = Visitor.objects(authorized=False).count()
    active_visitors = Visitor.objects(exit_time__exists=False).count()
    return render_template(
        'index.html',
        total_detections=total_detections,
        authorized_visitors=authorized_visitors,
        unauthorized_visitors=unauthorized_visitors,
        active_visitors=active_visitors
    )

@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('index'))
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        user = User.objects(username=username).first()
        if user:
            if user.failed_login_attempts >= 5 and (datetime.utcnow() - user.last_failed_login).total_seconds() < 300:
                flash('Account temporarily locked. Please try again later.')
                return render_template('login.html')
            if user.check_password(password):
                user.failed_login_attempts = 0
                user.last_login = datetime.utcnow()
                user.save()
                login_user(user)
                return redirect(url_for('index'))
            else:
                user.failed_login_attempts += 1
                user.last_failed_login = datetime.utcnow()
                user.save()
                if user.failed_login_attempts >= 5:
                    flash('Too many failed attempts. Account locked for 5 minutes.')
                else:
                    flash('Invalid username or password')
        else:
            flash('Invalid username or password')
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))

def preprocess_image(image):
    try:
        gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
        filtered = cv2.bilateralFilter(gray, 11, 17, 17)
        thresh = cv2.adaptiveThreshold(filtered, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY, 11, 2)
        kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (3, 3))
        morph = cv2.morphologyEx(thresh, cv2.MORPH_CLOSE, kernel)
        return morph
    except Exception as e:
        print(f"Error in image preprocessing: {str(e)}")
        return image

def enhance_plate_image(plate_image):
    try:
        gray = cv2.cvtColor(plate_image, cv2.COLOR_BGR2GRAY)
        clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8,8))
        enhanced = clahe.apply(gray)
        denoised = cv2.fastNlMeansDenoising(enhanced)
        enhanced_bgr = cv2.cvtColor(denoised, cv2.COLOR_GRAY2BGR)
        return enhanced_bgr
    except Exception as e:
        print(f"Error in plate enhancement: {str(e)}")
        return plate_image

def process_license_plate(image_path, plate_detector_model, ocr_model):
    try:
        img = cv2.imread(image_path)
        if img is None:
            return None, None, False, "Failed to load image file.", 0.0
        results = plate_detector_model(img)
        if len(results) == 0 or not results[0].boxes or len(results[0].boxes.data) == 0:
            return None, None, False, "No license plate detected.", 0.0
        box = results[0].boxes.data[0].tolist()
        x1, y1, x2, y2 = [int(i) for i in box[0:4]]
        plate_roi = img[y1:y2, x1:x2]
        if plate_roi.size == 0:
            return None, None, False, "Could not crop license plate area.", 0.0
        plate_roi = cv2.resize(plate_roi, (220, 220))
        temp_plate_filename = f"plate_{os.path.basename(image_path)}"
        temp_plate_path = os.path.join(TEMP_DIR, temp_plate_filename)
        cv2.imwrite(temp_plate_path, plate_roi)
        ocr_results = ocr_model(temp_plate_path)
        predicted_label = ''
        detected_chars_count = 0
        for result in ocr_results:
            if result.boxes and len(result.boxes.data) > 0:
                sorted_boxes = sorted(result.boxes.data, key=lambda box: box[0])
                for box in sorted_boxes:
                    x1_char, y1_char, x2_char, y2_char, conf, cls = box[:6]
                    class_label = CLASS_LABELS_MAPPING.get(int(cls.item()), 'Unknown')
                    color = COLORS[len(predicted_label) % len(COLORS)]
                    plate_roi = cv2.rectangle(plate_roi, (int(x1_char), int(y1_char)), (int(x2_char), int(y2_char)), color, 2)
                    plate_roi = draw_arabic_text(plate_roi, class_label, (int(x1_char), int(y1_char - 40)), color)
                    predicted_label += class_label
                    detected_chars_count += 1
        if os.path.exists(temp_plate_path):
            os.remove(temp_plate_path)
        if detected_chars_count == 0:
            return None, None, False, "No characters detected on the license plate.", 0.0
        predicted_label = reverse_arabic(predicted_label)
        plate_confidence = float(box[4]) if len(box) > 4 else 0.0
        return plate_roi, predicted_label, True, None, plate_confidence
    except Exception as e:
        print(f"Unexpected error during license plate processing: {str(e)}")
        return None, None, False, f"An unexpected error occurred: {str(e)}", 0.0

@app.route('/process_plate', methods=['POST'])
@login_required
def process_plate():
    if plate_detection_model is None or ocr_model is None:
        return jsonify({'error': 'Models not loaded properly'}), 500
    if 'image' not in request.files:
        return jsonify({'error': 'No image provided'}), 400
    image = request.files['image']
    if image.filename == '':
        return jsonify({'error': 'No selected file'}), 400
    if not image.filename.lower().endswith(tuple(ALLOWED_EXTENSIONS)):
        return jsonify({'error': f'Unsupported file type. Allowed types: {", ".join(ALLOWED_EXTENSIONS)}'}), 400
    os.makedirs(TEMP_DIR, exist_ok=True)
    temp_path = os.path.join(TEMP_DIR, image.filename)
    try:
        image.save(temp_path)
        processed_img, plate_text, success, error_msg, plate_confidence = process_license_plate(temp_path, plate_detection_model, ocr_model)
        if not success:
            if os.path.exists(temp_path):
                os.remove(temp_path)
            return jsonify({'success': False, 'message': error_msg, 'status': 'error'})
        _, buffer = cv2.imencode('.jpg', processed_img)
        processed_img_base64 = base64.b64encode(buffer).decode('utf-8')
        visitor = Visitor.objects(license_plate=plate_text).first()
        status = 'authorized' if visitor else 'unauthorized'
        visitor_name = visitor.name if visitor else None
        if visitor and not visitor.entry_datetime_utc:
            try:
                visitor.entry_datetime_utc = datetime.utcnow()
                visitor.save()
                app.logger.info(f"Automatically set entry time for visitor {visitor.name}")
            except Exception as update_error:
                app.logger.error(f"Error setting entry time for visitor {visitor.name}: {str(update_error)}")
        visitor_info = None
        if visitor:
            entry_time_str = None
            if visitor.entry_time:
                if isinstance(visitor.entry_time, str):
                    entry_time_str = visitor.entry_time
                elif isinstance(visitor.entry_time, datetime):
                    entry_time_str = utc_to_cairo(visitor.entry_time)
            exit_time_str = None
            if visitor.exit_time:
                if isinstance(visitor.exit_time, str):
                    exit_time_str = visitor.exit_time
                elif isinstance(visitor.exit_time, datetime):
                    exit_time_str = utc_to_cairo(visitor.exit_time)
            visitor_info = {
                'name': visitor.name,
                'visitor_id': str(visitor.visitor_id),
                'license_plate': visitor.license_plate,
                'entry_time': entry_time_str,
                'exit_time': exit_time_str,
                'authorized': visitor.authorized,
                'status': visitor.status,
                'responsible_department': visitor.responsible_department,
                'general_department': visitor.general_department
            }
        try:
            detection_record = DetectionResult(
                plate_number=plate_text,
                confidence=plate_confidence,
                status=status,
                visitor_name=visitor_name,
                processed_by=current_user.id,
                original_image=open(temp_path, 'rb').read(),
                processed_image=buffer.tobytes(),
                timestamp=datetime.utcnow()
            )
            detection_record.save()
        except Exception as db_error:
            print(f"Error saving detection record: {str(db_error)}")
            pass
        display_confidence = float(plate_confidence) if isinstance(plate_confidence, (int, float)) else 0.0
        detection_timestamp = detection_record.timestamp if detection_record else None
        return jsonify({
            'success': True,
            'plate_text': plate_text,
            'status': status,
            'visitor_name': visitor_name,
            'processed_image': processed_img_base64,
            'confidence': display_confidence,
            'visitor_id': str(visitor.id) if visitor else None,
            'visitor_info': visitor_info,
            'detection_timestamp': utc_to_cairo(detection_timestamp) if detection_timestamp else None
        })
    except Exception as e:
        print(f"Error processing image: {str(e)}")
        return jsonify({'error': str(e)}), 500
    finally:
        try:
            if os.path.exists(temp_path):
                os.remove(temp_path)
        except Exception as e:
            print(f"Error cleaning up temporary files: {str(e)}")

@app.route('/visitors')
@login_required
def visitors():
    visitors = Visitor.objects.all()
    return render_template('visitors.html', visitors=visitors)

@app.route('/reports')
@login_required
def reports():
    total_visitors = Visitor.objects.count()
    authorized_visitors = Visitor.objects(authorized=True).count()
    return render_template('reports.html', total_visitors=total_visitors, authorized_visitors=authorized_visitors)

@app.route('/users')
@login_required
def users():
    if not current_user.is_admin:
        flash('Access denied. Admin privileges required.')
        return redirect(url_for('index'))
    users = User.objects.all()
    return render_template('users.html', users=users)

@app.route('/api/users', methods=['POST'])
@login_required
def create_user():
    if not current_user.is_admin:
        return jsonify({'success': False, 'message': 'Access denied'}), 403
    username = request.form.get('username')
    password = request.form.get('password')
    is_admin = request.form.get('is_admin') == 'true'
    if not username or not password:
        return jsonify({'success': False, 'message': 'Username and password are required'}), 400
    if len(username) < 3:
        return jsonify({'success': False, 'message': 'Username must be at least 3 characters long'}), 400
    if len(password) < 6:
        return jsonify({'success': False, 'message': 'Password must be at least 6 characters long'}), 400
    if User.objects(username=username).first():
        return jsonify({'success': False, 'message': 'Username already exists'}), 400
    user = User(username=username, is_admin=is_admin)
    user.set_password(password)
    user.save()
    app.logger.info(f"User created by {current_user.username}: {username} (Admin: {is_admin})")
    return jsonify({'success': True})

@app.route('/api/users/<string:user_id>', methods=['DELETE'])
@login_required
def delete_user(user_id):
    if not current_user.is_admin:
        return jsonify({'success': False, 'message': 'Access denied'}), 403
    try:
        user = User.objects(id=ObjectId(user_id)).first()
        if not user:
            return jsonify({'success': False, 'message': 'User not found'}), 404
        if str(user.id) == str(current_user.id):
            return jsonify({'success': False, 'message': 'Cannot delete your own account'}), 400
        user.delete()
        app.logger.info(f"User deleted by {current_user.username}: {user.username} (ID: {user_id})")
        return jsonify({'success': True})
    except InvalidId:
        return jsonify({'success': False, 'message': 'Invalid user ID format'}), 400
    except Exception as e:
        app.logger.error(f"Error deleting user {user_id}: {str(e)}")
        return jsonify({'success': False, 'message': f'Error deleting user: {str(e)}'}), 500

@app.route('/api/users/<string:user_id>', methods=['PUT'])
@login_required
def update_user(user_id):
    if not current_user.is_admin:
        return jsonify({'success': False, 'message': 'Access denied'}), 403
    try:
        user = User.objects(id=ObjectId(user_id)).first()
        if not user:
            return jsonify({'success': False, 'message': 'User not found'}), 404
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'message': 'No data provided'}), 400
        if 'username' in data:
            if not data['username']:
                return jsonify({'success': False, 'message': 'Username cannot be empty'}), 400
            if len(data['username']) < 3:
                return jsonify({'success': False, 'message': 'Username must be at least 3 characters long'}), 400
            if User.objects(username=data['username'], id__ne=ObjectId(user_id)).first():
                return jsonify({'success': False, 'message': 'Username already exists'}), 400
            user.username = data['username']
        if 'password' in data and data['password']:
            if len(data['password']) < 6:
                return jsonify({'success': False, 'message': 'Password must be at least 6 characters long'}), 400
            user.set_password(data['password'])
        if 'is_admin' in data:
            user.is_admin = data['is_admin']
        user.save()
        app.logger.info(f"User updated by {current_user.username}: {user.username} (ID: {user_id}, Admin: {user.is_admin})")
        return jsonify({'success': True, 'message': 'User updated successfully'})
    except InvalidId:
        return jsonify({'success': False, 'message': 'Invalid user ID format'}), 400
    except Exception as e:
        app.logger.error(f"Error updating user {user_id}: {str(e)}")
        return jsonify({'success': False, 'message': f'Error updating user: {str(e)}'}), 500

@app.route('/api/users/<string:user_id>', methods=['GET'])
@login_required
def get_user(user_id):
    if not current_user.is_admin:
        return jsonify({'success': False, 'message': 'Access denied'}), 403
    try:
        user = User.objects(id=ObjectId(user_id)).first()
        if not user:
            return jsonify({'success': False, 'message': 'User not found'}), 404
        return jsonify({
            'success': True,
            'user': {
                'id': str(user.id),
                'username': user.username,
                'is_admin': user.is_admin
            }
        })
    except InvalidId:
        return jsonify({'success': False, 'message': 'Invalid user ID format'}), 400
    except Exception as e:
        app.logger.error(f"Error fetching user {user_id}: {str(e)}")
        return jsonify({'success': False, 'message': f'Error fetching user: {str(e)}'}), 500

@app.route('/api/users/<string:user_id>/activities')
@login_required
def get_user_activities(user_id):
    if not current_user.is_admin:
        return jsonify({'success': False, 'message': 'Access denied'}), 403
    try:
        user = User.objects(id=ObjectId(user_id)).first()
        if not user:
            return jsonify({'success': False, 'message': 'User not found'}), 404
        activities = UserActivity.objects(user=user).order_by('-timestamp').limit(50)
        return jsonify({
            'success': True,
            'activities': [{
                'action': activity.action,
                'details': activity.details,
                'ip_address': activity.ip_address,
                'timestamp': utc_to_cairo(activity.timestamp)
            } for activity in activities]
        })
    except InvalidId:
        return jsonify({'success': False, 'message': 'Invalid user ID format'}), 400
    except Exception as e:
        app.logger.error(f"Error fetching activities for user {user_id}: {str(e)}")
        return jsonify({'success': False, 'message': f'Error fetching activities: {str(e)}'}), 500

@app.before_request
def before_request():
    if current_user.is_authenticated:
        current_user.last_login = datetime.utcnow()
        current_user.save()

@app.after_request
def after_request(response):
    if current_user.is_authenticated:
        if request.endpoint and request.endpoint != 'static':
            activity = UserActivity(
                user=current_user,
                action=request.endpoint,
                details=f"{request.method} {request.path}",
                ip_address=request.remote_addr,
                timestamp=datetime.utcnow()
            )
            activity.save()
    return response

@app.route('/settings')
@login_required
def settings():
    if not current_user.is_admin:
        flash('Access denied. Admin privileges required.')
        return redirect(url_for('index'))
    settings = SystemSettings.objects.all()
    return render_template('settings.html', settings=settings)

@app.route('/api/settings', methods=['GET'])
@login_required
def get_settings():
    if not current_user.is_admin:
        return jsonify({'success': False, 'message': 'Access denied'}), 403
    settings = SystemSettings.objects.all()
    return jsonify({
        'success': True,
        'settings': [{
            'id': setting.id,
            'key': setting.key,
            'value': setting.value,
            'description': setting.description,
            'updated_at': utc_to_cairo(setting.updated_at)
        } for setting in settings]
    })

@app.route('/api/settings/<string:setting_id>', methods=['GET'])
@login_required
def get_setting(setting_id):
    if not current_user.is_admin:
        return jsonify({'success': False, 'message': 'Access denied'}), 403
    try:
        setting = SystemSettings.objects(id=ObjectId(setting_id)).first()
        if not setting:
            return jsonify({'success': False, 'message': 'Setting not found'}), 404
        return jsonify({
            'success': True,
            'setting': {
                'id': str(setting.id),
                'key': setting.key,
                'value': setting.value,
                'description': setting.description,
                'updated_at': utc_to_cairo(setting.updated_at)
            }
        })
    except (InvalidId, Exception) as e:
        app.logger.error(f"Error fetching setting {setting_id}: {str(e)}")
        return jsonify({'success': False, 'message': f'Error fetching setting: {str(e)}'}), 500

@app.route('/api/settings/<string:setting_id>', methods=['PUT'])
@login_required
def update_setting(setting_id):
    if not current_user.is_admin:
        return jsonify({'success': False, 'message': 'Access denied'}), 403
    try:
        setting = SystemSettings.objects(id=ObjectId(setting_id)).first()
        if not setting:
            return jsonify({'success': False, 'message': 'Setting not found'}), 404
        data = request.get_json()
        if 'value' in data:
            setting.value = data['value']
            setting.updated_by = current_user.id
        setting.save()
        return jsonify({'success': True})
    except (InvalidId, Exception) as e:
        app.logger.error(f"Error updating setting {setting_id}: {str(e)}")
        return jsonify({'success': False, 'message': f'Error updating setting: {str(e)}'}), 500

@app.route('/api/settings/<string:setting_id>', methods=['DELETE'])
@login_required
def delete_setting(setting_id):
    if not current_user.is_admin:
        return jsonify({'success': False, 'message': 'Access denied'}), 403
    try:
        setting = SystemSettings.objects(id=ObjectId(setting_id)).first()
        if not setting:
            return jsonify({'success': False, 'message': 'Setting not found'}), 404
        setting.delete()
        return jsonify({'success': True})
    except (InvalidId, Exception) as e:
        app.logger.error(f"Error deleting setting {setting_id}: {str(e)}")
        return jsonify({'success': False, 'message': f'Error deleting setting: {str(e)}'}), 500

@app.route('/detection_history')
@login_required
def detection_history():
    page = request.args.get('page', 1, type=int)
    per_page = 10
    skip = (page - 1) * per_page
    total = DetectionResult.objects.count()
    detections = DetectionResult.objects.order_by('-timestamp')[skip:skip + per_page]
    total_pages = (total + per_page - 1) // per_page
    return render_template(
        'detection_history.html',
        detections=detections,
        page=page,
        per_page=per_page,
        total=total,
        total_pages=total_pages,
        max=max,
        min=min
    )

@app.route('/api/detection/<string:detection_id>')
@login_required
def get_detection(detection_id):
    try:
        detection = DetectionResult.objects(id=detection_id).first()
        if not detection:
            return jsonify({'success': False, 'message': 'Detection record not found'}), 404
        processed_by_username = detection.processed_by.username if detection.processed_by else None
        visitor = Visitor.objects(license_plate=detection.plate_number).first() if detection.plate_number else None
        visitor_info = {}
        if visitor:
            entry_time_str = None
            if visitor.entry_time:
                if isinstance(visitor.entry_time, str):
                    entry_time_str = visitor.entry_time
                elif isinstance(visitor.entry_time, datetime):
                    entry_time_str = utc_to_cairo(visitor.entry_time)
            exit_time_str = None
            if visitor.exit_time:
                if isinstance(visitor.exit_time, str):
                    exit_time_str = visitor.exit_time
                elif isinstance(visitor.exit_time, datetime):
                    exit_time_str = utc_to_cairo(visitor.exit_time)
            visitor_info = {
                'name': visitor.name if visitor.name else 'N/A',
                'visitor_id': str(visitor.visitor_id) if visitor and visitor.visitor_id else 'N/A',
                'license_plate': visitor.license_plate if visitor else detection.plate_number,
                'entry_time': entry_time_str if entry_time_str else 'N/A',
                'exit_time': exit_time_str if exit_time_str else 'N/A',
                'department': visitor.responsible_department if visitor and visitor.responsible_department else 'N/A'
            }
        else:
            visitor_info = {
                'name': detection.visitor_name if detection.visitor_name else 'N/A',
                'visitor_id': 'N/A',
                'license_plate': detection.plate_number,
                'entry_time': 'N/A',
                'exit_time': 'N/A',
                'department': 'N/A'
            }
        timestamp_str = None
        if detection.timestamp:
            if isinstance(detection.timestamp, str):
                timestamp_str = detection.timestamp
            elif isinstance(detection.timestamp, datetime):
                timestamp_str = utc_to_cairo(detection.timestamp)
        else:
            app.logger.warning(f"Detection {detection_id} has no timestamp")
        return jsonify({
            'success': True,
            'id': str(detection.id),
            'timestamp': timestamp_str,
            'plate_number': detection.plate_number,
            'confidence': float(detection.confidence) if detection.confidence is not None else 0.0,
            'status': detection.status,
            'visitor_name': detection.visitor_name,
            'processed_by': processed_by_username,
            'original_image': base64.b64encode(detection.original_image).decode('utf-8') if detection.original_image else '',
            'processed_image': base64.b64encode(detection.processed_image).decode('utf-8') if detection.processed_image else '',
            'visitor_info': visitor_info
        })
    except Exception as e:
        app.logger.error(f"Error fetching detection {detection_id}: {str(e)}")
        return jsonify({'success': False, 'message': f'Error fetching detection details: {str(e)}'}), 500

@app.route('/api/detections/clear', methods=['POST'])
@login_required
def clear_detection_history():
    if not current_user.is_admin:
        return jsonify({'success': False, 'message': 'Access denied'}), 403
    try:
        DetectionResult.objects.delete()
        app.logger.info(f"Detection history cleared by user: {current_user.username} at {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')}")
        return jsonify({'success': True, 'message': 'Detection history cleared successfully'})
    except Exception as e:
        app.logger.error(f"Error clearing detection history: {str(e)}", exc_info=True)
        return jsonify({'success': False, 'message': f'Failed to clear detection history: {str(e)}'}), 500

@app.route('/manage_visitors', methods=['GET', 'POST'])
@login_required
def manage_visitors():
    if request.method == 'POST':
        action = request.form.get('action')
        visitor_id = request.form.get('visitor_id')
        if action == 'clear':
            if not current_user.is_admin:
                flash('Access denied. Admin privileges required.', 'error')
                return redirect(url_for('manage_visitors'))
            try:
                Visitor.objects.delete()
                flash('All visitors cleared successfully', 'success')
                app.logger.info(f"All visitors cleared by user: {current_user.username}")
            except Exception as e:
                flash(f'Failed to clear visitors: {str(e)}', 'error')
                app.logger.error(f"Error clearing visitors: {str(e)}", exc_info=True)
            return redirect(url_for('manage_visitors'))
        if not visitor_id:
            flash('Visitor ID is required', 'error')
            return redirect(url_for('manage_visitors'))
        try:
            visitor = Visitor.objects(visitor_id=int(visitor_id)).first()
        except ValueError:
            flash('Invalid visitor ID', 'error')
            return redirect(url_for('manage_visitors'))
        if not visitor:
            flash('Visitor not found', 'error')
            return redirect(url_for('manage_visitors'))
        if action == 'authorize':
            visitor.authorized = True
            visitor.status = 'authorized'
            flash('Visitor authorized successfully', 'success')
        elif action == 'unauthorize':
            visitor.authorized = False
            visitor.status = 'unauthorized'
            flash('Visitor unauthorized successfully', 'success')
        elif action == 'delete':
            visitor.delete()
            flash('Visitor deleted successfully', 'success')
        elif action == 'update':
            name = request.form.get('name')
            visitor_code = request.form.get('visitor_code')
            license_plate = request.form.get('license_plate')
            if name:
                visitor.name = name
            if visitor_code:
                visitor.visitor_code = visitor_code
                if visitor_code.startswith('V'):
                    try:
                        visitor.visitor_id = int(visitor_code.lstrip('V'))
                    except ValueError:
                        flash('Invalid visitor code format', 'error')
                        return redirect(url_for('manage_visitors'))
            if license_plate:
                visitor.license_plate = license_plate
            flash('Visitor information updated successfully', 'success')
        visitor.save()
        return redirect(url_for('manage_visitors'))
    search_query = request.args.get('search_query', '')
    query = Visitor.objects
    if search_query:
        query = query.filter(Q(name__icontains=search_query) | Q(visitor_code__icontains=search_query) | Q(license_plate__icontains=search_query))
    visitors = query.order_by('-entry_datetime_utc').all()
    return render_template('manage_visitors.html', visitors=visitors, search_query=search_query)

@app.route('/api/visitors', methods=['POST'])
@login_required
def add_visitor():
    name = request.form.get('name')
    visitor_code = request.form.get('visitor_code')
    license_plate = request.form.get('license_plate')
    status = request.form.get('status', 'pending')
    responsible_department = request.form.get('responsible_department')
    general_department = request.form.get('general_department')
    app.logger.info(f"Received POST to /api/visitors with data: name={name}, visitor_code={visitor_code}, license_plate={license_plate}, status={status}, responsible_department={responsible_department}, general_department={general_department}")
    if not all([name, visitor_code, license_plate]):
        app.logger.warning("Missing required fields in visitor creation request")
        return jsonify({'success': False, 'message': 'All fields (name, visitor_code, license_plate) are required'}), 400
    existing_visitor = Visitor.objects(Q(visitor_code=visitor_code) | Q(license_plate=license_plate)).first()
    if existing_visitor:
        app.logger.warning(f"Duplicate visitor detected - Code: {visitor_code}, Plate: {license_plate}, Existing ID: {existing_visitor.id}")
        return jsonify({'success': False, 'message': 'Visitor with this code or license plate already exists'}), 400
    try:
        visitor = Visitor(
            name=name,
            visitor_code=visitor_code,
            license_plate=license_plate,
            status=status,
            responsible_department=responsible_department,
            general_department=general_department,
            authorized=False,
            entry_datetime_utc=datetime.utcnow()
        )
        if visitor_code.startswith('V'):
            try:
                visitor.visitor_id = int(visitor_code.lstrip('V'))
            except ValueError:
                app.logger.warning(f"Invalid visitor_code format: {visitor_code}")
                return jsonify({'success': False, 'message': 'Invalid visitor code format. Must be in format V<number>'}), 400
        visitor.save()
        app.logger.info(f"Visitor added successfully - ID: {visitor.id}, Name: {name}, Code: {visitor_code}")
        return jsonify({
            'success': True,
            'message': 'Visitor added successfully',
            'visitor': {
                'id': str(visitor.id),
                'name': visitor.name,
                'visitor_id': str(visitor.visitor_id) if visitor.visitor_id else None,
                'visitor_code': visitor.visitor_code,
                'license_plate': visitor.license_plate,
                'status': visitor.status,
                'responsible_department': visitor.responsible_department,
                'general_department': visitor.general_department
            }
        })
    except Exception as e:
        app.logger.error(f"Error saving visitor to database: {str(e)}", exc_info=True)
        return jsonify({'success': False, 'message': f'Failed to add visitor: {str(e)}'}), 500

@app.route('/api/visitors/<string:visitor_id>', methods=['GET'])
@login_required
def get_visitor(visitor_id):
    if not visitor_id.isdigit():
        return jsonify({'success': False, 'message': 'Invalid visitor ID'}), 400
    visitor = Visitor.objects(visitor_id=int(visitor_id)).first()
    if not visitor:
        return jsonify({'success': False, 'message': 'Visitor not found'}), 404
    return jsonify({
        'id': str(visitor.id),
        'name': visitor.name,
        'visitor_id': str(visitor.visitor_id),
        'visitor_code': visitor.visitor_code,
        'license_plate': visitor.license_plate,
        'entry_time': utc_to_cairo(visitor.entry_time),
        'entry_date': visitor.entry_date,
        'exit_time': utc_to_cairo(visitor.exit_time) if visitor.exit_time else None,
        'authorized': visitor.authorized,
        'status': visitor.status,
        'responsible_department': visitor.responsible_department if visitor.responsible_department else None,
        'general_department': visitor.general_department if visitor.general_department else None
    })

@app.route('/api/visitors/<string:visitor_id>', methods=['PUT'])
@login_required
def update_visitor(visitor_id):
    if not visitor_id.isdigit():
        return jsonify({'success': False, 'message': 'Invalid visitor ID'}), 400
    visitor = Visitor.objects(visitor_id=int(visitor_id)).first()
    if not visitor:
        return jsonify({'success': False, 'message': 'Visitor not found'}), 404
    data = request.get_json()
    if 'name' in data:
        visitor.name = data['name']
    if 'visitor_code' in data:
        visitor.visitor_code = data['visitor_code']
        if data['visitor_code'].startswith('V'):
            try:
                visitor.visitor_id = int(data['visitor_code'].lstrip('V'))
            except ValueError:
                return jsonify({'success': False, 'message': 'Invalid visitor_code format'}), 400
    if 'license_plate' in data:
        visitor.license_plate = data['license_plate']
    if 'authorized' in data:
        visitor.authorized = data['authorized']
        visitor.status = 'authorized' if data['authorized'] else 'unauthorized'
    if 'exit_time' in data:
        visitor.exit_time = datetime.strptime(data['exit_time'], '%Y-%m-%d %H:%M:%S')
    entry_date_str = data.get('entry_date')
    entry_time_str = data.get('entry_time')
    if entry_date_str and entry_time_str:
        try:
            entry_datetime_str = f'{entry_date_str} {entry_time_str}'
            visitor.entry_datetime_utc = datetime.strptime(entry_datetime_str, '%Y-%m-%d %H:%M:%S')
        except ValueError:
            return jsonify({'success': False, 'message': 'Invalid date or time format for entry time'}), 400
    visitor.save()
    return jsonify({'success': True, 'message': 'Visitor updated successfully'})
@app.route('/api/visitors/<string:visitor_id>', methods=['DELETE'])
@login_required
def delete_visitor(visitor_id):
    if not visitor_id.isdigit():
        return jsonify({'success': False, 'message': 'Invalid visitor ID'}), 400
    visitor = Visitor.objects(visitor_id=int(visitor_id)).first()
    if not visitor:
        return jsonify({'success': False, 'message': 'Visitor not found'}), 404
    visitor.delete()
    return jsonify({'success': True, 'message': 'Visitor deleted successfully'})

@app.route('/api/visitors/clear', methods=['POST'])
@login_required
def clear_visitors():
    if not current_user.is_admin:
        return jsonify({'success': False, 'message': 'Access denied. Admin privileges required.'}), 403
    try:
        Visitor.objects.delete()
        app.logger.info(f"All visitors cleared by user: {current_user.username} at {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')}")
        return jsonify({'success': True, 'message': 'All visitors cleared successfully'})
    except Exception as e:
        app.logger.error(f"Error clearing visitors: {str(e)}", exc_info=True)
        return jsonify({'success': False, 'message': f'Failed to clear visitors: {str(e)}'}), 500

@app.route('/uploaded_images')
@login_required
def uploaded_images():
    try:
        detections = DetectionResult.objects.order_by('-timestamp').all()
        return render_template('uploaded_images.html', detections=detections)
    except Exception as e:
        app.logger.error(f"Error retrieving detections for uploaded_images: {str(e)}", exc_info=True)
        flash('An error occurred while loading images. Please try again.', 'error')
        return render_template('uploaded_images.html', detections=[])

@app.route('/api/images/<int:detection_id>')
@login_required
def get_detection_image(detection_id):
    detection = DetectionResult.objects(id=detection_id).first()
    return jsonify({
        'original_image': base64.b64encode(detection.original_image).decode('utf-8'),
        'processed_image': base64.b64encode(detection.processed_image).decode('utf-8'),
        'plate_number': detection.plate_number,
        'status': detection.status,
        'timestamp': utc_to_cairo(detection.timestamp),
        'visitor_name': detection.visitor_name
    })

@app.route('/api/export/detections')
@login_required
def export_detections():
    try:
        status = request.args.get('status', 'all')
        date_filter = request.args.get('date_filter', 'all')
        format = request.args.get('format', 'csv')
        app.logger.info(f"Export request - Status: {status}, Date: {date_filter}, Format: {format}")
        query = DetectionResult.objects
        if status != 'all':
            query = query.filter(status=status)
        if date_filter != 'all':
            now = datetime.utcnow()
            if date_filter == 'today':
                query = query.filter(timestamp__gte=now.replace(hour=0, minute=0, second=0))
            elif date_filter == 'week':
                query = query.filter(timestamp__gte=now - timedelta(days=7))
            elif date_filter == 'month':
                query = query.filter(timestamp__gte=now - timedelta(days=30))
        results = query.order_by('-timestamp')
        if not results:
            app.logger.warning("No data found for the selected filters")
            return jsonify({'error': 'No data found for the selected filters'}), 404
        app.logger.info(f"Found {results.count()} records to export")
        if format == 'csv':
            output = io.StringIO()
            writer = csv.writer(output)
            writer.writerow(['Timestamp', 'Plate Number', 'Status', 'Confidence', 'Visitor Name', 'Processed By'])
            for result in results:
                processed_by_username = result.processed_by.username if result.processed_by else 'N/A'
                writer.writerow([
                    utc_to_cairo(result.timestamp),
                    result.plate_number,
                    result.status,
                    f"{result.confidence:.2f}%" if result.confidence else 'N/A',
                    result.visitor_name or 'N/A',
                    processed_by_username
                ])
            output.seek(0)
            return Response(
                output,
                mimetype='text/csv',
                headers={
                    'Content-Disposition': f'attachment; filename=detections_export_{datetime.utcnow().strftime("%Y%m%d")}.csv'
                }
            )
        elif format == 'excel':
            wb = Workbook()
            ws = wb.active
            ws.title = "Detection History"
            headers = ['Timestamp', 'Plate Number', 'Status', 'Confidence', 'Visitor Name', 'Processed By']
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            for row, result in enumerate(results, 2):
                processed_by_username = result.processed_by.username if result.processed_by else 'N/A'
                ws.cell(row=row, column=1, value=utc_to_cairo(result.timestamp))
                ws.cell(row=row, column=2, value=result.plate_number)
                ws.cell(row=row, column=3, value=result.status)
                ws.cell(row=row, column=4, value=f"{result.confidence:.2f}%" if result.confidence else 'N/A')
                ws.cell(row=row, column=5, value=result.visitor_name or 'N/A')
                ws.cell(row=row, column=6, value=processed_by_username)
            excel_file = io.BytesIO()
            wb.save(excel_file)
            excel_file.seek(0)
            return Response(
                excel_file,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                headers={
                    'Content-Disposition': f'attachment; filename=detections_export_{datetime.utcnow().strftime("%Y%m%d")}.xlsx'
                }
            )
        else:
            app.logger.error(f"Unsupported export format: {format}")
            return jsonify({'error': 'Unsupported export format'}), 400
    except Exception as e:
        app.logger.error(f"Export error: {str(e)}", exc_info=True)
        return jsonify({'error': f'Failed to export data: {str(e)}'}), 500

@app.route('/api/activities')
@login_required
def get_activities():
    try:
        activities = DetectionResult.objects.order_by('-timestamp').limit(10)
        activity_list = []
        for activity in activities:
            details = f"Detected plate {activity.plate_number} - {activity.status}"
            activity_list.append({
                'action': 'detect_plate',
                'details': details,
                'timestamp': utc_to_cairo(activity.timestamp)
            })
        return jsonify({
            'success': True,
            'activities': activity_list
        })
    except Exception as e:
        app.logger.error(f"Error fetching activities: {str(e)}")
        return jsonify({
            'success': False,
            'error': 'Failed to fetch facilities'
        }), 500

@app.route('/process_video', methods=['POST'])
@login_required
def process_video():
    app.logger.info(f"Received video processing request from user: {current_user.username}")
    if plate_detection_model is None or ocr_model is None:
        app.logger.error("Models not loaded properly for video processing")
        return jsonify({'error': 'Models not loaded properly'}), 500

    if 'video' not in request.files:
        app.logger.warning("No video file provided in request")
        return jsonify({'error': 'No video file provided'}), 400

    video = request.files['video']
    if video.filename == '':
        app.logger.warning("No selected file filename provided")
        return jsonify({'error': 'No selected file'}), 400

    file_ext = video.filename.rsplit('.', 1)[1].lower() if '.' in video.filename else ''
    if file_ext not in ALLOWED_VIDEO_EXTENSIONS:
        app.logger.warning(f"Unsupported video file extension: {file_ext}")
        return jsonify({'error': f'Unsupported file type. Allowed types: {", ".join(ALLOWED_VIDEO_EXTENSIONS)}'}), 400

    # Check file size
    video.seek(0, os.SEEK_END)
    file_size = video.tell()
    video.seek(0)
    if file_size > MAX_VIDEO_SIZE:
        app.logger.warning(f"Video file size ({file_size} bytes) exceeds limit ({MAX_VIDEO_SIZE} bytes)")
        return jsonify({'error': f'Video file size exceeds limit of {MAX_VIDEO_SIZE / 1024 / 1024:.2f} MB'}), 400

    # Save temporary file
    os.makedirs(TEMP_DIR, exist_ok=True)
    temp_path = os.path.join(TEMP_DIR, video.filename)
    output_path = os.path.join(TEMP_DIR, f"processed_{video.filename}")
    
    try:
        video.save(temp_path)
        app.logger.info(f"Processing video: {video.filename}")
        
        cap = cv2.VideoCapture(temp_path)
        if not cap.isOpened():
            raise Exception("Could not open video file")

        # Get video properties
        fps = cap.get(cv2.CAP_PROP_FPS)
        frame_width = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
        frame_height = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
        total_frames = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))

        # Create video writer
        fourcc = cv2.VideoWriter_fourcc(*'mp4v')
        out = cv2.VideoWriter(output_path, fourcc, fps, (frame_width, frame_height))
        
        frame_count = 0
        processed_frames = 0
        best_detections = {}
        last_frame = None

        while True:
            ret, frame = cap.read()
            if not ret:
                break

            frame_count += 1
            if frame_count % VIDEO_FRAME_SKIP == 0:
                detection_data, processed_img, current_frame = process_frame_and_detect(
                    frame, frame_count, plate_detection_model, ocr_model, None, last_frame
                )
                
                if detection_data:
                    plate_text = detection_data['plate_text']
                    if (plate_text not in best_detections or 
                        detection_data['confidence'] > best_detections[plate_text]['confidence']):
                        best_detections[plate_text] = detection_data
                        try:
                            socketio.emit('live_detection', detection_data)
                            app.logger.debug(f"Emitted detection for plate: {plate_text}")
                        except Exception as emit_error:
                            app.logger.error(f"Error emitting detection: {str(emit_error)}")

                # Write frame to output video
                if processed_img is not None:
                    out.write(processed_img)
                else:
                    out.write(frame)

                processed_frames += 1
                last_frame = current_frame

        cap.release()
        out.release()

        # Read processed video
        if os.path.exists(output_path):
            with open(output_path, 'rb') as video_file:
                video_base64 = base64.b64encode(video_file.read()).decode('utf-8')
        else:
            video_base64 = None

        # Clean up temporary files
        cleanup_paths = [temp_path, output_path]
        for path in cleanup_paths:
            if os.path.exists(path):
                os.remove(path)

        return jsonify({
            'success': True,
            'message': 'Video processing completed',
            'total_frames': total_frames,
            'processed_frames': processed_frames,
            'detections': list(best_detections.values()),
            'processed_video': video_base64
        })
    except Exception as e:
        app.logger.error(f"Error processing video: {str(e)}", exc_info=True)
        # Clean up temporary files
        cleanup_paths = [temp_path, output_path]
        for path in cleanup_paths:
            if os.path.exists(path):
                try:
                    os.remove(path)
                except Exception:
                    pass
        return jsonify({'error': f'An error occurred during video processing: {str(e)}'}), 500

camera = None

def process_frame_and_detect(frame, frame_count, plate_detection_model, ocr_model, last_detection_time=None, last_frame=None):
    """Helper function to process frame and emit detection results for both video and live camera"""
    # Check frame similarity if we have a previous frame
    if last_frame is not None:
        frame_diff = cv2.absdiff(frame, last_frame)
        change_percent = np.sum(frame_diff > 30) / frame_diff.size * 100
        if change_percent < 1.0:  # Less than 1% change
            return None, None, frame

    temp_frame_path = os.path.join(TEMP_DIR, f"frame_{frame_count}.jpg")
    try:
        cv2.imwrite(temp_frame_path, frame)
        processed_img, plate_text, processing_success, error_msg, plate_confidence = process_license_plate(
            temp_frame_path,
            plate_detection_model,
            ocr_model
        )

        if processing_success:
            current_time = datetime.utcnow()
            
            # Skip duplicate detections if tracking last detection time
            if last_detection_time and plate_text in last_detection_time:
                time_since_last = (current_time - last_detection_time[plate_text]).total_seconds()
                if time_since_last < VIDEO_DUPLICATE_WINDOW:
                    return None, None, frame

            # Look up visitor information
            visitor = Visitor.objects(license_plate=plate_text).first()
            status = 'authorized' if visitor else 'unauthorized'
            
            # Format visitor information
            visitor_info = None
            if visitor:
                visitor_info = {
                    'name': visitor.name,
                    'license_plate': visitor.license_plate,
                    'entry_time': utc_to_cairo(visitor.entry_time) if visitor.entry_time else 'N/A',
                    'exit_time': utc_to_cairo(visitor.exit_time) if visitor.exit_time else 'N/A',
                    'responsible_department': visitor.responsible_department or 'غير محدد',
                    'general_department': visitor.general_department or 'غير محدد'
                }
            
            # Save detection to database
            try:
                # Encode images to binary format
                _, img_encoded = cv2.imencode('.jpg', frame)
                original_image = img_encoded.tobytes()
                
                _, processed_encoded = cv2.imencode('.jpg', processed_img)
                processed_image = processed_encoded.tobytes()
                
                # Create and save detection result
                detection_result = DetectionResult(
                    plate_number=plate_text,
                    confidence=plate_confidence,
                    status=status,
                    timestamp=current_time,
                    visitor_name=visitor_info['name'] if visitor_info else None,
                    original_image=original_image,
                    processed_image=processed_image,
                    processed_by=current_user if current_user else None
                )
                detection_result.save()
                app.logger.info(f"Saved detection result for plate {plate_text}")
            except Exception as db_error:
                app.logger.error(f"Error saving detection result: {str(db_error)}")

            # Create detection data for response
            detection_data = {
                'plate_text': plate_text,
                'status': status,
                'confidence': plate_confidence,
                'timestamp': utc_to_cairo(current_time),
                'visitor_info': visitor_info,
                'frame': frame_count,
                'processed_image': base64.b64encode(cv2.imencode('.jpg', processed_img)[1]).decode('utf-8') if processed_img is not None else None,
                'original_image': base64.b64encode(cv2.imencode('.jpg', frame)[1]).decode('utf-8')
            }

            if last_detection_time is not None:
                last_detection_time[plate_text] = current_time

            return detection_data, processed_img, frame

        return None, processed_img, frame

    except Exception as e:
        app.logger.error(f"Error in frame processing: {str(e)}", exc_info=True)
        return None, None, frame
    finally:
        if os.path.exists(temp_frame_path):
            try:
                os.remove(temp_frame_path)
            except Exception:
                pass

def generate_frames():
    global camera
    if camera is None or not camera.isOpened():
        camera = cv2.VideoCapture(0)
        if not camera.isOpened():
            app.logger.error("Could not open camera.")
            return
    app.logger.info("Starting camera frame generation...")
    last_detection_time = {}
    frame_count = 0
    
    while True:
        success, frame = camera.read()
        if not success:
            app.logger.warning("Failed to read frame from camera stream.")
            break
            
        frame_count += 1
        if frame_count % VIDEO_FRAME_SKIP == 0:
            detection_data, processed_img = process_frame_and_detect(
                frame, frame_count, plate_detection_model, ocr_model, last_detection_time
            )
            
            if detection_data:
                try:
                    socketio.emit('live_detection', detection_data)
                    app.logger.debug(f"Emitted live_detection for plate: {detection_data['plate_text']}")
                except Exception as emit_error:
                    app.logger.error(f"Error emitting live detection data: {str(emit_error)}", exc_info=True)
        
        try:
            ret, buffer = cv2.imencode('.jpg', frame)
            if not ret:
                app.logger.error("Failed to encode frame to JPEG for streaming.")
                continue
            frame_bytes = buffer.tobytes()
            yield (b'--frame\r\n'
                   b'Content-Type: image/jpeg\r\n\r\n' + frame_bytes + b'\r\n')
        except Exception as encode_error:
            app.logger.error(f"Error encoding frame {frame_count} for streaming: {str(encode_error)}", exc_info=True)
            
    if camera and camera.isOpened():
        camera.release()
        app.logger.info("Camera released.")

@app.route('/video_feed')
@login_required 
def video_feed():
    app.logger.info("Video feed requested.")
    if plate_detection_model is None or ocr_model is None:
        app.logger.error("Models not loaded for video feed.")
        return Response("Models not loaded properly", status=500)
    return Response(generate_frames(), mimetype='multipart/x-mixed-replace; boundary=frame')

@socketio.on('connect')
def handle_connect():
    app.logger.info("Client connected to SocketIO.")

@socketio.on('disconnect')
def handle_disconnect():
    app.logger.info("Client disconnected from SocketIO.")

@app.route('/api/settings', methods=['PUT'])
@login_required
def update_detection_settings():
    if not current_user.is_admin:
        return jsonify({'success': False, 'message': 'Access denied'}), 403
    data = request.get_json()
    try:
        # Update each setting by key
        if 'confidence_threshold' in data:
            SystemSettings.objects(key='confidence_threshold').update_one(set__value=str(data['confidence_threshold']))
        if 'processing_mode' in data:
            SystemSettings.objects(key='processing_mode').update_one(set__value=data['processing_mode'])
        if 'auto_process' in data:
            SystemSettings.objects(key='auto_process').update_one(set__value=str(data['auto_process']).lower())
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

if __name__ == '__main__':
    socketio.run(app, debug=True)  